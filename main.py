import ollama
import easyocr
import regex
import time
import re
from transformers import DonutProcessor, VisionEncoderDecoderModel
import numpy as np
from PIL import Image
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import json

model_name = "mychen76/invoice-and-receipts_donut_v1"

processor = DonutProcessor.from_pretrained(model_name)
model = VisionEncoderDecoderModel.from_pretrained(model_name)
model.eval()

def increment_letter(letter):
    if letter == 'Z':
        raise ValueError("Cannot increment beyond 'Z'")
    return chr(ord(letter) + 1)

class ModelLoader:
    _instance = None

    def __new__(cls):
        if cls._instance is None:
            cls._instance = super(ModelLoader, cls).__new__(cls)
            cls._instance._initialize_model()
        return cls._instance

    def _initialize_model(self):
        self.reader = easyocr.Reader(['en'])  # Initialize your EasyOCR reader or other models here

    def get_model(self):
        return self.reader

def get_easyocr_reader():
    model_loader = ModelLoader()
    return model_loader.get_model()


def extract_float_from_string(s: str) -> float:
    pattern = r'[-+]?\d*\.\d+|[-+]?\d+'

    clean_string = re.sub(r'[^\d\.\-+]', ' ', s) 

    matches = re.findall(pattern, clean_string)

    if matches:
        return float(matches[0])
    else:
        return None

def replace_except_chars(s, chars_to_keep, replacement_char):
    pattern = f'[^{re.escape(chars_to_keep)}]'
    return re.sub(pattern, replacement_char, s)

def getNumber(stringThatIsAlMostAnumber):
    stringThatIsAlMostAnumber = stringThatIsAlMostAnumber.strip()
    stringThatIsAlMostAnumber = stringThatIsAlMostAnumber.replace(',', '.')
    chars_to_keep = "1234567890."
    return replace_except_chars(stringThatIsAlMostAnumber, chars_to_keep, "")

def resultIsTotal(text):
    formattedText = text.strip().lower()

    possibleResults = [
        # English
        "total", "balance", "net total", "payment", "amount",
        
        # French
        "total", "solde", "total net", "paiement", "montant",

        # Spanish
        "total", "saldo", "total neto", "pago", "cantidad",

        # German
        "gesamt", "saldo", "nettobetrag", "zahlung", "betrag",

        # Italian
        "totale", "saldo", "totale netto", "pagamento", "importo",

        # Portuguese
        "total", "saldo", "total líquido", "pagamento", "quantia",

        # Dutch
        "totaal", "saldo", "netto totaal", "betaling", "bedrag",

        # Swedish
        "total", "saldo", "nettobelopp", "betalning", "belopp"
    ]

    for result in possibleResults:
        max_errors = max(1, len(result) // 3) 
        if regex.search(f'({result}){{e<={max_errors}}}', formattedText):
            return True

    return False

def findTextOnTheSameLine(thetotalword, word):
    """
    Checks if the bounding box of a text is on the same line with another bounding box while considering a
    dynamically chosen vertical offset error.
    """
    bbox = np.array(thetotalword, dtype=np.int32)
    firsty = bbox[0][1]
    secondy = bbox[3][1]

    bbox = np.array(word, dtype=np.int32)
    newfirsty = bbox[0][1]
    newsecondy = bbox[3][1]

    c = (secondy - firsty) / 2

    return abs(firsty-newfirsty) < c or abs(secondy-newsecondy) < c

def imageProcessorWithEasyocr(img_path):
    """
    Process a receipt by using only EasyOCR. No other post-processing.

    Details: Extract text with EasyOcr, find the "TOTAL" label and look for a price tag on the same line with the label.
    """
    start_time = time.time()
    reader = get_easyocr_reader() 
    results = reader.readtext(img_path, width_ths=10000, min_size=5, link_threshold=0.1)

    listOfTotals = [result for result in results if resultIsTotal(result[1])]

    listOfPrices = []
    for result in listOfTotals:
        for words in results:
            if findTextOnTheSameLine(result[0], words[0]) and result != words:
                number = extract_float_from_string(words[1])
                if number:
                    listOfPrices.append(words)

    for item in listOfTotals:
        number = extract_float_from_string(item[1])
        if number:
            listOfPrices.append(item)
    end_time = time.time()

    try:
        #trebuie sa extragem tot ce inseaman 
        maxim, conf = -1, -1
        for item in listOfPrices:
            if extract_float_from_string(item[1]) and extract_float_from_string(item[1]) > maxim:
                maxim = extract_float_from_string(item[1])
                conf = int(item[2])
        return maxim, end_time - start_time
    except Exception as e:
        return "0.0", end_time - start_time

def imageProcessorWithEasyOcrAndLLM(img_path, model_name, prompt):
    """
    Process a receipt by using EasyOCR along with a LLM for post-processing.

    Details: Extract text with EasyOCR, pass the result line by line to an LLM and ask it about the total price.

    :param model_name: name of the ollama model to be used
    :param prompt: custom instructions for the model
    """
    start_time = time.time()
    reader = get_easyocr_reader()
    results = reader.readtext(img_path, width_ths=10000, min_size=5, link_threshold=0.1)

    concatenated_results = ""
    for r in results:
        concatenated_results += f"{r[1]}\n"

    response = ollama.chat(model=model_name, messages=[
        {
            'role': 'user',
            'content': f'{prompt} . Context: "{concatenated_results}"'
        },
    ])

    total_price_str = getNumber(response['message']['content'])

    end_time = time.time()
    return extract_float_from_string(total_price_str), end_time - start_time

def getInformationFromReceipt(img_path):
    start_time = time.time()
    try:
        image = Image.open(img_path).convert("RGB")
        pixel_values = processor(image, return_tensors="pt").pixel_values
        generated_ids = model.generate(pixel_values)
        processing =  processor.batch_decode(generated_ids, skip_special_tokens=True)
        output_text = processing[0]
        total_matches = re.findall(r'<s_total>(.*?)</s_total>', output_text)

        if total_matches:
            total_text = total_matches[0]
            
            regex = r'^[^0-9]*([0-9]+\.[0-9]{2}).*$'
            pattern = re.compile(regex)
            
            match = pattern.search(total_text)

            end_time = time.time()
            if match:
                return match.group(1), end_time - start_time
        
        return  getNumber(total_matches[0]), end_time - start_time
    except Exception as e:
        print(f"Error processing receipt: {e}")
        return ""

def mainFunctionForStatistics():
    json_file_path = './test/metadata.jsonl'
    workbook = load_workbook('statictisc.xlsx')
    sheet = workbook.active

    final_data = {}
    index = 0
    with open(json_file_path, 'r') as file:
        for line in file:
            try:
                data = json.loads(line.strip())  
                final_data[index] = data
                index += 1
                #pretty_data = json.dumps(data, indent=4) 
                
            except json.JSONDecodeError:
                print("Error decoding a line in the JSONL file")
    
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    yello_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

     #stat of avrg 

    model1_time = 0
    model2_time = 0
    model3_time = 0
    no_of_pics = 0

    model1_corr = 0
    model1_wrong = 0
    model1_emtpy = 0
    model2_corr = 0
    model2_wrong = 0
    model2_emtpy = 0
    model3_corr = 0
    model3_wrong = 0
    model3_emtpy = 0
    #end of avrg

    row = 2
    number_of_runs = 100
    for key in final_data:
        values = final_data[key]

        file_name = values['file_name']
        file_path = './test/' + file_name

        final_price = values['ground_truth']['gt_parse']['total']
        final_price = getNumber(final_price)

        regex = r'^[^0-9]*([0-9]+\.[0-9]{2}).*$'
        pattern = re.compile(regex)
        match = pattern.search(final_price)
        if match:
            final_price = match.group(1)

        # using EasyOCR + LLM
        result1, time1= imageProcessorWithEasyOcrAndLLM(
            file_path,
            model_name="mistral",
            prompt=(
                f'You are given raw text data of a receipt. Extract and return only the total amount of money to be paid by the customer as a number (e.g., 123.45).')
        )

        result2, time2 = getInformationFromReceipt(file_path)

        result3, time3 = imageProcessorWithEasyocr(file_path)

        no_of_pics += 1
        model1_time += time1
        model2_time += time2
        model3_time += time3

        column = 'A'

        print("---------------------------------------------------------")
        print(file_path + " -> " + final_price + " and the result are -> " + str(result1) + " in " + str("{:.2f}".format(time1)) + " seconds > " + str(result2)+ " in " + str("{:.2f}".format(time2)) + " seconds > " + str(result3)+ " in " + str("{:.2f}".format(time3)) + " seconds ")
        print("---------------------------------------------------------")
        #now that we have the info we need to start to fill the statisctis

        val1 = extract_float_from_string(str(result1))
        val2 = extract_float_from_string(str(result2))
        val3 = extract_float_from_string(str(result3))
        compareto = extract_float_from_string(str(final_price))
        sheet[column+str(row)].value = file_name

        column = increment_letter(column)
        sheet[column+str(row)].value = getNumber(str(final_price))

        column = increment_letter(column)
        sheet[column+str(row)].value = getNumber(str(result1))

        column = increment_letter(column)
        sheet[column+str(row)].value = "{:.2f}".format(time1)

        column = increment_letter(column)
        if val1 == compareto:
            sheet[column+str(row)].fill = green_fill
            model1_corr += 1
        elif str(val1) in str(compareto) and str(val1):
            sheet[column+str(row)].fill = yello_fill
        else:
            sheet[column+str(row)].fill = red_fill
            model1_wrong += 1

        column = increment_letter(column)
        sheet[column+str(row)].value = getNumber(str(result2))

        column = increment_letter(column)
        sheet[column+str(row)].value = "{:.2f}".format(time2)

        column = increment_letter(column)
        if val2 == compareto:
            sheet[column+str(row)].fill = green_fill
            model2_corr += 1
        elif str(val2) in str(compareto) and str(val2):
            sheet[column+str(row)].fill = yello_fill
        else:
            sheet[column+str(row)].fill = red_fill
            model2_wrong += 1

        column = increment_letter(column)
        sheet[column+str(row)].value = getNumber(str(result3))

        column = increment_letter(column)
        sheet[column+str(row)].value = "{:.2f}".format(time3)

        column = increment_letter(column)
        if str(val3) == "-1":
            model3_emtpy += 1
        if val3 == compareto:
            model3_corr += 1
            sheet[column+str(row)].fill = green_fill
        elif str(val3) in str(compareto) and str(val3):
            sheet[column+str(row)].fill = yello_fill
        else:
            sheet[column+str(row)].fill = red_fill
            model3_wrong += 1

        row += 1

        if row >= number_of_runs:
            break


    model1_time /= no_of_pics
    model2_time /= no_of_pics
    model3_time /= no_of_pics
    sheet['M2'].value = f"{model1_time:.2f}"
    sheet['M3'].value = f"{model2_time:.2f}"
    sheet['M4'].value = f"{model3_time:.2f}"
    sheet['N2'].value = f"{no_of_pics}"
    sheet['O2'].value = f"{model1_corr}"
    sheet['P2'].value = f"{model1_wrong}"
    sheet['Q2'].value = f"{model1_emtpy}"
    sheet['R2'].value = f"{model1_corr / no_of_pics:.2f}"
    sheet['S2'].value = f"{model1_corr / (model1_corr + model1_wrong):.2f}"
    sheet['N3'].value = f"{no_of_pics}"
    sheet['O3'].value = f"{model2_corr}"
    sheet['P3'].value = f"{model2_wrong}"
    sheet['Q3'].value = f"{model2_emtpy}"
    sheet['R3'].value = f"{model2_corr / no_of_pics:.2f}"
    sheet['S3'].value = f"{model2_corr / (model2_corr + model2_wrong):.2f}"
    sheet['N4'].value = f"{no_of_pics}"
    sheet['O4'].value = f"{model3_corr}"
    sheet['P4'].value = f"{model3_wrong}"
    sheet['Q4'].value = f"{model3_emtpy}"
    sheet['R4'].value = f"{model3_corr / no_of_pics:.2f}"
    sheet['S4'].value = f"{model3_corr / (model3_corr + model3_wrong):.2f}"
    workbook.save('statictisc.xlsx')
    


mainFunctionForStatistics()
